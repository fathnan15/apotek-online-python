from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, NoAlertPresentException
from openpyxl import load_workbook
import time

# Load Excel file
file_path = "/home/fathnan/Asterix/list_sep.xlsx"
wb = load_workbook(file_path)
sheet = wb["sep_web_driver"]

# Connect to existing Chrome session
# google-chrome --remote-debugging-port=9222 --user-data-dir="$HOME/.chrome_debug"
options = webdriver.ChromeOptions()
options.debugger_address = "localhost:9222"
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 5)

def reset_form():
    """Click Reset and wait until SEP input is cleared."""
    driver.find_element(By.ID, "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_BtnReset_CD").click()
    WebDriverWait(driver, 3).until(
        lambda d: d.find_element(By.ID, "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_TxtREFASALSJP_I")
                   .get_attribute("value").strip() == ""
    )

i = 2  # Start from row 2

while i <= sheet.max_row:
    sep_no     = sheet[f'B{i}'].value
    receipt_no = sheet[f'C{i}'].value
    status_cd  = sheet[f'D{i}'].value

    # skip empty rows or already-normal
    if sep_no is None and receipt_no is None:
        i += 1
        continue
    if status_cd == "normal":
        i += 1
        continue

    try:
        # Input SEP and click Cari
        sep_input = driver.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_TxtREFASALSJP_I")
        sep_input.clear()
        sep_input.send_keys(sep_no)
        driver.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_BtnCariSEP_CD").click()

        # --- IMMEDIATE ALERT CHECK ---
        try:
            alert = WebDriverWait(driver, 2).until(EC.alert_is_present())
            alert_text = alert.text.strip()
            alert.accept()

            # Log error in Excel
            sheet[f'D{i}'].value = "error"
            sheet[f'E{i}'].value = f"Cari alert: {alert_text}"
            wb.save(file_path)

            # reset and skip to next
            reset_form()
            print(f"Row {i} {sep_no} – Cari-alert (“{alert_text}”) handled; moving on.")
            i += 1
            continue

        except TimeoutException:
            # no alert after Cari → proceed
            pass

        # Wait for Nomor Kartu field to populate
        wait.until(lambda d: d.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_txtNOKAPST_I")
            .get_attribute("value").strip() != ""
        )

        # Fill Jenis Resep and No Resep
        driver.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_cboJnsObat_I") \
              .send_keys("Obat Kronis Blm Stabil")
        driver.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_txtNoResep_I") \
              .send_keys(receipt_no)

        # Click Simpan
        driver.find_element(By.ID, 
            "ctl00_ctl00_ASPxSplitter1_Content_ContentSplitter_MainContent_BtnSimpan_CD").click()

        # Handle the save-alert
        try:
            alert = WebDriverWait(driver, 5).until(EC.alert_is_present())
            alert_text = alert.text.strip()
            alert.accept()

            if "Simpan Berhasil" in alert_text:
                sheet[f'D{i}'].value = "normal"
                print(f"Row {i} {sep_no} – Saved successfully.")
            else:
                sheet[f'D{i}'].value = "error"
                sheet[f'E{i}'].value = alert_text
                print(f"Row {i} {sep_no} – Save error: {alert_text}")
                reset_form()

            wb.save(file_path)

        except TimeoutException:
            print(f"Row {i} {sep_no} – No alert after Simpan.")

    except Exception as e:
        # unexpected failure: log and continue
        print(f"Row {i} {sep_no} – Unexpected exception: {e}")

    i += 1

print("Process Completed")
