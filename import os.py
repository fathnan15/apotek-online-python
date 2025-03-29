import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
root_folder = r"/home/fathnan/Downloads/Asterix"
echo_folder = os.path.join(root_folder, "documents/echocardiogram")
excel_file = r"/home/fathnan/Downloads/Asterix/list_needfile_2.xlsx"
report_file = r"/home/fathnan/Desktop/asterix/report.txt"

# Load Excel
wb = load_workbook(excel_file)
sheet = wb['inject']

# Define color fills for Excel status
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Success
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Copied
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Not Found
gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Moved

# Define keyword mapping for different categories
file_categories = {
    "D": ["echocardiogram", "lv ef", "eko", "echo"],  # Echocardiogram-related
    "E": ["lab", "laboratorium"],  # Lab-related
    "F": ["resep"],  # Prescription-related
    "G": ["billing", "bil", "tagihan"]  # Billing-related
}

# Dictionary to track where each sep_num has already been moved
sep_locations = {}

# Open report file
with open(report_file, "w", encoding="utf-8") as report:
    report.write("Missing Files Report\n")
    report.write("=" * 40 + "\n")

    # Iterate through Excel rows
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        sep_num, dest_folder, card_num = row[:3]  # Read values from Excel

        if not sep_num or not dest_folder:
            print(f"Skipping row {i} due to missing data.")
            continue  # Skip if missing data

        full_dest_folder = os.path.join(root_folder, dest_folder)
        found_folder = None

        # Check if the destination folder already contains a folder with sep_num
        folder_exists = False
        for folder_name in os.listdir(full_dest_folder):
            if sep_num in folder_name:
                folder_exists = True
                found_folder = os.path.join(full_dest_folder, folder_name)
                sep_locations[sep_num] = found_folder  # Track the existing location
                break

        if folder_exists:
            print(f"Folder containing {sep_num} already exists in {full_dest_folder}.")
            sheet[f'A{i}'].fill = green_fill  # Mark as success since the folder already exists
        else:
            # Check if sep_num has already been moved to another destination
            if sep_num in sep_locations:
                source_folder = sep_locations[sep_num]
                shutil.copytree(source_folder, os.path.join(full_dest_folder, os.path.basename(source_folder)))
                found_folder = os.path.join(full_dest_folder, os.path.basename(source_folder))
                print(f"Copied {source_folder} to {full_dest_folder}")
                sheet[f'A{i}'].fill = yellow_fill  # Mark as copied
                sheet[f'H{i}'] = f"Copied from {source_folder}"
            else:
                # Search for the folder in the root folder (excluding the destination folder)
                found = False
                for root, dirs, files in os.walk(root_folder):
                    if root == full_dest_folder:
                        continue
                    for dir_name in dirs:
                        if sep_num in dir_name:
                            source_folder = os.path.join(root, dir_name)
                            shutil.move(source_folder, os.path.join(full_dest_folder, dir_name))
                            found_folder = os.path.join(full_dest_folder, dir_name)
                            sep_locations[sep_num] = found_folder  # Track the new location
                            print(f"Moved {source_folder} to {full_dest_folder}")
                            sheet[f'A{i}'].fill = gray_fill  # Mark as moved
                            sheet[f'H{i}'] = f"Moved from {source_folder}"
                            found = True
                            break
                    if found:
                        break

                if not found:
                    report.write(f"❌ No folder found for SEP: {sep_num}\n")
                    sheet[f'A{i}'].fill = red_fill  # Mark as failure
                    print(f"❌ No folder found for SEP: {sep_num}")
                    continue

        # Check for echocardiogram file
        echo_file = None
        print(f"Searching for echocardiogram file for card number: {card_num}")
        for file in os.listdir(echo_folder):
            print(f"Checking file: {file}")
            if f"_{card_num}_" in file and file.lower().endswith("_echo.pdf"):
                echo_file = os.path.join(echo_folder, file)
                break

        if echo_file:
            # Copy the echocardiogram file to the found folder
            shutil.copy(echo_file, found_folder)
            report.write(f"📄 Copied {os.path.basename(echo_file)} → {found_folder}\n")
            sheet[f'C{i}'].fill = green_fill  # Mark as success
            print(f"📄 Copied {os.path.basename(echo_file)} → {found_folder}")
        else:
            report.write(f"❌ No echocardiogram file found for card number: {card_num}\n")
            print(f"❌ No echocardiogram file found for card number: {card_num}")

        # Check for required documents inside the SEP folder
        found_categories = {key: False for key in file_categories}  # Track found categories

        for file_name in os.listdir(found_folder):
            lower_file = file_name.lower()
            for col, keywords in file_categories.items():
                if any(keyword in lower_file for keyword in keywords):
                    sheet[f'{col}{i}'] = "V"  # Mark as found
                    found_categories[col] = True

        # Log missing categories
        missing_categories = [col for col, found in found_categories.items() if not found]
        if missing_categories:
            missing_labels = [key for key, col in file_categories.items() if key in missing_categories]
            report.write(f"⚠️ Folder for SEP {sep_num} is **missing**: {', '.join(missing_labels)}\n")
            print(f"⚠️ Folder for SEP {sep_num} is **missing**: {', '.join(missing_labels)}")

print(f"✅ Process completed. Check {report_file} for missing files.")
wb.save(excel_file)