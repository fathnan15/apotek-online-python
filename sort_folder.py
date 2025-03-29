import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the sheet
wb = load_workbook('/home/fathnan/Desktop/asterix/list_needfile.xlsx')
sheet = wb['Sheet1']

# Paths (Update these as needed)
root_folder = r"/home/fathnan/Downloads/Asterix"

# Define the red fill for failed cells, green fill for existing folders, and yellow fill for moved folders
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

i = 2

while i < len(sheet['A'])+1:
    sep_num = sheet['A' + str(i)].value
    dest_folder = os.path.join(root_folder, sheet['B' + str(i)].value)

    # Check if the destination folder already contains a folder with sep_num
    folder_exists = False
    for folder_name in os.listdir(dest_folder):
        if sep_num in folder_name:
            folder_exists = True
            break

    if folder_exists:
        print(f"Folder containing {sep_num} already exists in {dest_folder}. Skipping...")
        sheet['A' + str(i)].fill = green_fill  # Mark as success since the folder already exists
    else:
        # Search for the folder in the root folder (excluding the destination folder)
        found = False
        for root, dirs, files in os.walk(root_folder):
            if root == dest_folder:
                continue
            for dir_name in dirs:
                if sep_num in dir_name:
                    source_folder = os.path.join(root, dir_name)
                    shutil.move(source_folder, os.path.join(dest_folder, dir_name))
                    print(f"Moved {source_folder} to {dest_folder}")
                    found = True
                    sheet['A' + str(i)].fill = yellow_fill  # Mark as moved
                    break
            if found:
                break

        if not found:
            print(f"Folder containing {sep_num} not found in {root_folder}")
            sheet['A' + str(i)].fill = red_fill  # Mark as failure

    i += 1

# Save the workbook with the updated cell colors
wb.save('/home/fathnan/Desktop/asterix/list_needfile.xlsx')