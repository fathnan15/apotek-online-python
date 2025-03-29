import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
root_folder = r"/home/fathnan/Downloads/Asterix"
echo_folder = os.path.join(root_folder, "documents/echocardiogram")
excel_file = r"/home/fathnan/Downloads/Asterix/list_needfile_2.xlsx"
report_file = r"/home/fathnan/Desktop/asterix/report.txt"

# Load Excel
wb = load_workbook(excel_file)
sheet = wb['inject']

# Define color fills for Excel status
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Success
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Moved
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Not Found

# Define keyword mapping for different categories
file_categories = {
    "D": ["echocardiogram", "lv ef", "eko", "echo"],  # Echocardiogram-related
    "E": ["lab", "laboratorium"],  # Lab-related
    "F": ["resep"],  # Prescription-related
    "G": ["billing", "bil", "tagihan"]  # Billing-related
}

# Open report file
with open(report_file, "w", encoding="utf-8") as report:
    report.write("Missing Files Report\n")
    report.write("=" * 40 + "\n")

    # Iterate through Excel rows
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        sep_num, dest_folder, card_num = row[:3]  # Read values from Excel

        if not sep_num or not dest_folder:
            print(f"Skipping row {i} due to missing data.")
            continue  # Skip if missing data

        full_dest_folder = os.path.join(root_folder, dest_folder)
        found_folder = None

        # Search for the folder containing SEP number
        print(f"Searching for SEP: {sep_num} in {root_folder}")
        for root, dirs, _ in os.walk(root_folder):
            if root == full_dest_folder:  # Skip destination folder
                continue
            for folder_name in dirs:
                print(f"Checking folder: {folder_name} in {root}")
                if sep_num in folder_name:
                    found_folder = os.path.join(root, folder_name)
                    print(f"Found folder: {found_folder}")
                    break
            if found_folder:
                break

        if not found_folder:
            report.write(f"❌ No folder found for SEP: {sep_num}\n")
            sheet[f'A{i}'].fill = red_fill  # Mark as failure
            print(f"❌ No folder found for SEP: {sep_num}")
            continue

        # Move/copy the folder to the destination
        new_folder_path = os.path.join(full_dest_folder, os.path.basename(found_folder))
        if not os.path.exists(new_folder_path):
            shutil.move(found_folder, new_folder_path)
            report.write(f"📂 Moved {found_folder} → {full_dest_folder}\n")
            sheet[f'A{i}'].fill = yellow_fill  # Mark as moved
            print(f"📂 Moved {found_folder} → {full_dest_folder}")
        else:
            report.write(f"⚠️ Folder for SEP {sep_num} already exists in {full_dest_folder}\n")
            print(f"⚠️ Folder for SEP {sep_num} already exists in {full_dest_folder}")

        # Check for echocardiogram file
        echo_file = None
        for file in os.listdir(echo_folder):
            if f"_{card_num}_" in file and file.endswith("_echo.pdf"):
                echo_file = os.path.join(echo_folder, file)
                break

        if echo_file:
            shutil.copy(echo_file, new_folder_path)
            report.write(f"📄 Copied {os.path.basename(echo_file)} → {new_folder_path}\n")
            sheet[f'C{i}'].fill = green_fill  # Mark as success
            print(f"📄 Copied {os.path.basename(echo_file)} → {new_folder_path}")

        # Check for required documents inside the SEP folder
        found_categories = {key: False for key in file_categories}  # Track found categories

        for file_name in os.listdir(new_folder_path):
            lower_file = file_name.lower()
            for col, keywords in file_categories.items():
                if any(keyword in lower_file for keyword in keywords):
                    sheet[f'{col}{i}'] = "V"  # Mark as found
                    found_categories[col] = True

        # Log missing categories
        missing_categories = [col for col, found in found_categories.items() if not found]
        if missing_categories:
            missing_labels = [key for key, col in file_categories.items() if key in missing_categories]
            report.write(f"⚠️ Folder for SEP {sep_num} is **missing**: {', '.join(missing_labels)}\n")
            print(f"⚠️ Folder for SEP {sep_num} is **missing**: {', '.join(missing_labels)}")

print(f"✅ Process completed. Check {report_file} for missing files.")
wb.save(excel_file)